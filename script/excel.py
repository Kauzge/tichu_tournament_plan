# pylint: skip-file
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
from openpyxl.utils import get_column_letter

def create_formula(column_shift:int, player_id:int, number_of_rounds:int, number_of_tables:int, number_of_players:int) -> str:
    # Create the formula for the total tichu or victory points of a player
    formula = "="
    first_occurence = False
    for table_id in range(number_of_tables):
        # Create the formula for each table
        for team in range(2):
            # Repeat it twice (for the two teams of each table)
            for seat in range(2):
                # Repeat it twice (for the two seats of each team)
                if first_occurence:
                    # Add a "+" between the formulas on every occurence except the first one
                    formula += "+"
                first_occurence = True
                # Concatenate the formulas for each of the two column per table
                formula += f"SUMIF(${get_column_letter((table_id*9)+3+seat+(team*4))}{number_of_players+6}:{get_column_letter((table_id*9)+3+seat+(team*4))}{number_of_players+number_of_rounds+5},$B{player_id+1},{get_column_letter((table_id*9)+5+column_shift+(team*4))}{number_of_players+6}:{get_column_letter((table_id*9)+5+column_shift+(team*4))}{number_of_players+number_of_rounds+5})"
    return formula

def create_player_table(players:list, number_of_rounds:int, number_of_tables:int, number_of_players:int, ws:Workbook) -> None:
    # Create the player table to insert the names of the players
    ws.cell(row=1, column=1).value = "Spieler"
    ws.cell(row=1, column=2).value = "Name"
    ws.cell(row=1, column=3).value = "TP gesamt"
    ws.cell(row=1, column=4).value = "SP gesamt"
    for player_id in players:
        # Add the names of the players
        ws.cell(row=player_id+1, column=1).value = player_id
        ws.cell(row=player_id+1, column=2).value = player_id
        # Add the formulas for the total tichu and victory points
        ws.cell(row=player_id+1, column=3).value = create_formula(0, player_id, number_of_rounds, number_of_tables, number_of_players)
        ws.cell(row=player_id+1, column=4).value = create_formula(1, player_id, number_of_rounds, number_of_tables, number_of_players)

def create_headers(number_of_players:int, number_of_tables:int, ws:Workbook) -> None:
    # Create the headers for the tournament plan
    offset_y = number_of_players + 4
    ws.cell(row=offset_y+1, column=1).value = "Runde"
    for table_id in range(number_of_tables):
        # Add headers for each table
        offset_x = (table_id*9) + 3
        table = ws.cell(row=offset_y, column=offset_x)
        table.value = f"Tisch {table_id+1}"
        ws.merge_cells(start_row=offset_y, start_column=offset_x, end_row=offset_y, end_column=offset_x+7)
        table.alignment = Alignment(horizontal='center')
        for team_id in range(1,3):
            # Add headers for team 1 and team 2
            team_offset = offset_x+(team_id*4)-4
            team = ws.cell(row=offset_y+1, column=team_offset)
            ws.merge_cells(start_row=offset_y+1, start_column=team_offset, end_row=offset_y+1, end_column=team_offset+1)
            team.value = f"Team {team_id}"
            team.alignment = Alignment(horizontal='center')
            # Add headers for tichu points and victory points
            tichu_points = ws.cell(row=offset_y+1, column=team_offset+2)
            victory_points = ws.cell(row=offset_y+1, column=team_offset+3)
            tichu_points.value = "TP"
            victory_points.value = "SP"

def create_tournament_table(tournament_plan:list, number_of_players:int, ws:Workbook) -> None:
    # Create the tournament table with the tournament plan
    for round_id,round in enumerate(tournament_plan):
        offset_y = round_id + number_of_players + 6
        ws.cell(row=offset_y, column=1).value = round_id+1
        for table_id,table in enumerate(round):
            offset_x = (table_id*9) + 2 
            for i in range(4):
                # Insert the players for each team
                player = ws.cell(row=offset_y, column=(i+1)+offset_x)
                player.value = f"=B{table[i]+1}"
                if i == 0:
                    player.border = Border(left=Side(border_style="thin",color='00000000'))
                if i == 1 or i == 3:
                    # Insert cells for tichu points and victory points for each team
                    tichu_points = ws.cell(row=offset_y, column=(i+1)+offset_x+1)
                    victory_points = ws.cell(row=offset_y, column=(i+1)+offset_x+2)
                    tichu_points.value = 0
                    victory_points.value = 0
                    victory_points.border = Border(right=Side(border_style="thin",color='00000000'))
                    offset_x += 2

def create_excel(players:list, tournament_plan:list, number_of_players:int, number_of_tables:int, number_of_rounds:int) -> None:
    # Create the excel file with the tournament plan
    today = datetime.today().strftime('%d_%m_%Y')
    wb = Workbook()
    ws = wb.active
    create_player_table(players, number_of_rounds, number_of_tables, number_of_players, ws)
    create_headers(number_of_players, number_of_tables, ws)
    create_tournament_table(tournament_plan, number_of_players, ws)
    wb.save(f'tichu_tournament_{today}.xlsx')
    
if __name__ == '__main__':
    players = [x for x in range(1, 9)]
    number_of_players = len(players)
    number_of_tables = number_of_players // 4
    number_of_rounds = 6
    tournament_plan =  [[[5, 7, 2, 6], [4, 1, 8, 3]],
                        [[3, 5, 8, 4], [1, 6, 2, 7]],
                        [[2, 1, 6, 5], [8, 7, 4, 3]],
                        [[5, 1, 6, 4], [3, 7, 8, 2]],
                        [[1, 8, 3, 2], [6, 7, 4, 5]],
                        [[7, 1, 2, 4], [5, 8, 3, 6]]]
    create_excel(players, tournament_plan, number_of_players, number_of_tables, number_of_rounds)
    
    